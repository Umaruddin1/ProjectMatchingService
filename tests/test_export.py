"""Tests for export service."""
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.export_service import ExportService


def test_create_export_workbook_returns_bytes_without_writing_file(monkeypatch, tmp_path):
    """Export generation should stay in memory and not write to disk."""
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        "app.services.export_service.datetime",
        type("FixedDateTime", (), {"now": staticmethod(lambda: datetime(2026, 4, 29, 12, 0, 0))}),
    )

    file_bytes, file_name = ExportService.create_export_workbook(
        reconciled_matches=[
            {
                "current_row_number": 2,
                "current_project": "Project A",
                "previous_row_number": 2,
                "previous_project": "Project A",
                "match_status": "matched",
                "requires_review": False,
                "current_values": {"opening_balance": 100, "additions": 50, "transfer": 25, "closing_balance": 125},
                "previous_values": {"opening_balance": 80, "additions": 40, "transfer": 20, "closing_balance": 100},
                "wip_impact": 25,
                "far_impact": 5,
            }
        ],
        unmatched_current_rows=[],
        unmatched_previous_rows=[],
        validation_issues=[],
        summary={"total_matched": 1, "total_unmatched_current": 0, "total_unmatched_previous": 0},
    )

    assert file_name == "reconciliation_export_20260429_120000.xlsx"
    assert isinstance(file_bytes, bytes)
    assert not Path(file_name).exists()

    workbook = load_workbook(BytesIO(file_bytes))
    assert workbook.sheetnames == ["Summary", "Matched Results"]
    assert workbook["Summary"]["A1"].value == "Summary"
    assert workbook["Matched Results"]["A2"].value == 2