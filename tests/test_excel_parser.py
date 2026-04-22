"""Tests for Excel parsing."""
import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook
from app.services.excel_parser_service import ExcelParserService
from app.core.exceptions import ParsingException


class TestExcelParser:
    """Test Excel parsing functionality."""
    
    @staticmethod
    def create_test_workbook():
        """Create a test Excel workbook."""
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Create current year sheet
        cy = wb.create_sheet("Current Year")
        cy.append(["Project Name", "As of 31 Mar", "Additions", "Transfer", "As on 31 Mar"])
        cy.append(["Project A", 100, 50, 25, 125])
        cy.append(["Project B", 200, 75, 50, 225])
        
        # Create previous year sheet
        py = wb.create_sheet("Previous Year")
        py.append(["Project Name", "Opening Balance", "Additions", "Transfer", "Closing Balance"])
        py.append(["Project A", 80, 40, 20, 100])
        py.append(["Project B", 150, 60, 40, 170])
        
        return wb
    
    def test_parse_valid_workbook(self):
        """Test parsing a valid workbook."""
        wb = self.create_test_workbook()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)
            
            try:
                current, previous, current_headers, previous_headers = (
                    ExcelParserService.parse_workbook(tmp_path)
                )
                
                assert len(current) == 2
                assert len(previous) == 2
                assert len(current_headers) == 5
                assert len(previous_headers) == 5
            finally:
                # Cleanup
                try:
                    Path(tmp_path).unlink()
                except:
                    pass
    
    def test_find_sheet_by_keywords(self):
        """Test finding sheet by keywords."""
        sheet_names = ["Current Year", "Previous Year"]
        
        current = ExcelParserService.find_sheet_by_keywords(
            sheet_names, {"current year", "cy"}
        )
        assert current == "Current Year"
        
        previous = ExcelParserService.find_sheet_by_keywords(
            sheet_names, {"previous year", "py"}
        )
        assert previous == "Previous Year"
    
    def test_find_sheet_not_found(self):
        """Test sheet not found."""
        sheet_names = ["Data", "Summary"]
        
        with pytest.raises(ParsingException):
            ExcelParserService.find_sheet_by_keywords(
                sheet_names, {"current year"}
            )
    
    def test_normalize_headers(self):
        """Test header normalization."""
        headers = ["Project Name", "As of 31 Mar", "OPENING BALANCE"]
        
        normalized = ExcelParserService.normalize_headers(headers)
        
        assert "project name" in normalized
        assert "as of 31 mar" in normalized
        assert "opening balance" in normalized
    
    def test_validate_headers_success(self):
        """Test successful header validation."""
        headers = ["Project Name", "Opening Balance", "Additions", "Transfer", "Closing Balance"]
        required = {"project name", "opening balance", "additions", "transfer", "closing balance"}
        
        mapping, issues = ExcelParserService.validate_and_extract_headers(
            headers, required, "previous_year"
        )
        
        assert len(mapping) >= len(required)
        assert len(issues) == 0
    
    def test_validate_headers_missing(self):
        """Test header validation with missing required field."""
        headers = ["Project Name", "Additions", "Transfer"]
        required = {"project name", "opening balance", "additions", "transfer", "closing balance"}
        
        with pytest.raises(Exception):  # ValidationException
            ExcelParserService.validate_and_extract_headers(
                headers, required, "previous_year"
            )
