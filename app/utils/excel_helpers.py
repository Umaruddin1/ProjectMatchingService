"""Excel helper utilities."""
import logging
from openpyxl import load_workbook
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


def load_excel_workbook(file_path: str):
    """
    Load an Excel workbook from file.
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        Workbook object from openpyxl
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        return workbook
    except Exception as e:
        logger.error(f"Failed to load workbook: {e}")
        raise


def get_sheet_names(workbook) -> List[str]:
    """Get all sheet names from workbook."""
    return workbook.sheetnames


def read_sheet_data(sheet) -> List[Dict[str, any]]:
    """
    Read data from a sheet into list of dicts.
    Robustly detect the header row as the first row with at least 2 non-empty cells.
    
    Args:
        sheet: Worksheet object
        
    Returns:
        List of row dicts, header list
    """
    rows = []
    headers = None
    header_row_idx = None
    
    # Find header row: first row with >=2 non-empty cells
    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        non_empty = [cell for cell in row if cell is not None and str(cell).strip() != '']
        if headers is None and len(non_empty) >= 2:
            headers = [h if h is not None else '' for h in row]
            header_row_idx = idx
            break
    if headers is None:
        raise ValueError("No header row found in sheet")
    
    # Read data rows after header
    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if idx <= header_row_idx:
            continue
        if not any(row):  # Skip empty rows
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
            else:
                row_dict[header] = None
        rows.append(row_dict)
    
    return rows, headers
