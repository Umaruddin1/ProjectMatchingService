"""Export service."""
import logging
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

logger = logging.getLogger(__name__)


class ExportService:
    """Generates Excel exports."""
    
    @staticmethod
    def create_export_workbook(
        reconciled_matches: List[Dict[str, Any]],
        unmatched_current_rows: List[Dict[str, Any]],
        unmatched_previous_rows: List[Dict[str, Any]],
        validation_issues: List[Dict[str, Any]],
        summary: Dict[str, Any],
    ) -> str:
        """
        Create and save an Excel workbook with reconciliation results.
        
        Args:
            reconciled_matches: List of reconciled matches
            unmatched_current_rows: Unmatched current year rows
            unmatched_previous_rows: Unmatched previous year rows
            validation_issues: Validation issues found
            summary: Summary statistics
            
        Returns:
            str: Path to created file
        """
        logger.info("Creating export workbook")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        summary_font = Font(bold=True)
        
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Create sheets
        ExportService._create_summary_sheet(wb, summary, header_fill, header_font, summary_fill, summary_font, border)
        ExportService._create_matched_sheet(wb, reconciled_matches, header_fill, header_font, border)
        ExportService._create_unmatched_sheet(wb, unmatched_current_rows, "Unmatched Current Year", header_fill, header_font, border)
        ExportService._create_unmatched_sheet(wb, unmatched_previous_rows, "Unmatched Previous Year", header_fill, header_font, border)
        
        if validation_issues:
            ExportService._create_validation_sheet(wb, validation_issues, header_fill, header_font, border)
        
        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"reconciliation_export_{timestamp}.xlsx"
        file_path = Path(file_name).absolute()
        
        wb.save(str(file_path))
        logger.info(f"Export saved: {file_path}")
        
        return str(file_path)
    
    @staticmethod
    def _create_summary_sheet(wb, summary, header_fill, header_font, summary_fill, summary_font, border):
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Headers
        ws["A1"] = "Summary"
        ws["A1"].font = Font(bold=True, size=12)
        
        row = 3
        for key, value in summary.items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            ws[f"A{row}"].font = summary_font
            ws[f"B{row}"] = value
            row += 1
        
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
    
    @staticmethod
    def _create_matched_sheet(wb, reconciled_matches, header_fill, header_font, border):
        """Create matched results sheet."""
        ws = wb.create_sheet("Matched Results")
        
        # Headers
        headers = [
            "Current Row",
            "Current Project",
            "Previous Row",
            "Previous Project",
            "Match Status",
            "Requires Review",
            "Current Opening",
            "Current Additions",
            "Current Transfer",
            "Current Closing",
            "Previous Opening",
            "Previous Additions",
            "Previous Transfer",
            "Previous Closing",
            "WIP Impact",
            "FAR Impact",
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data rows
        for row_idx, match in enumerate(reconciled_matches, 2):
            current_vals = match.get("current_values", {})
            previous_vals = match.get("previous_values", {})
            
            # Handle both "project_name" and "current_project"/"previous_project" field names
            current_project = match.get("current_project") or match.get("project_name", "")
            previous_project = match.get("previous_project", "")
            match_status = match.get("match_status", "matched")
            requires_review = match.get("requires_review", False)
            
            data = [
                match.get("current_row_number"),
                current_project,
                match.get("previous_row_number"),
                previous_project,
                match_status,
                requires_review,
                current_vals.get("opening_balance", 0) or current_vals.get("as_of_31_mar", 0),
                current_vals.get("additions", 0),
                current_vals.get("transfer", 0),
                current_vals.get("closing_balance", 0) or current_vals.get("as_on_31_mar", 0),
                previous_vals.get("opening_balance", 0),
                previous_vals.get("additions", 0),
                previous_vals.get("transfer", 0),
                previous_vals.get("closing_balance", 0),
                match.get("wip_impact", 0),
                match.get("far_impact", 0),
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"
        
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    
    @staticmethod
    def _create_unmatched_sheet(wb, unmatched_rows, sheet_name, header_fill, header_font, border):
        """Create unmatched rows sheet."""
        if not unmatched_rows:
            return
        
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = [
            "Row Number",
            "Project Name",
            "Opening Balance",
            "Additions",
            "Transfer",
            "Closing Balance",
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        # Data rows
        for row_idx, row in enumerate(unmatched_rows, 2):
            values_dict = row.get("values", {})
            
            data = [
                row.get("row_number"),
                row.get("project_name"),
                values_dict.get("opening_balance", 0),
                values_dict.get("additions", 0),
                values_dict.get("transfer", 0),
                values_dict.get("closing_balance", 0),
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"
        
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15
    
    @staticmethod
    def _create_validation_sheet(wb, validation_issues, header_fill, header_font, border):
        """Create validation issues sheet."""
        ws = wb.create_sheet("Validation Issues")
        
        # Headers
        headers = ["Row Number", "Project Name", "Issue Type", "Description"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        # Data rows
        for row_idx, issue in enumerate(validation_issues, 2):
            data = [
                issue.get("row_number"),
                issue.get("project_name", ""),
                issue.get("issue_type", ""),
                issue.get("description", ""),
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 50
